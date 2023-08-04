import openpyxl as opx


def set_database(filename):
    wb = opx.load_workbook(filename=filename)
    ws = wb.active

def get_product(barcode:int):
    wb = opx.load_workbook(filename="database/slives.xlsx")
    ws = wb.active
    for bc in range(2,ws.row_max+1):
        if int(ws["B"+str(bc)]) == barcode:
            return True
        else:
            return False
    
def get_paquet(barcode:int):
    wb = opx.load_workbook(filename="database/paquet.xlsx")
    ws = wb.active
    for bc in range(2,ws.row_max+1):
        if int(ws["B"+str(bc)]) == barcode:
            return True
        else:
            return False
